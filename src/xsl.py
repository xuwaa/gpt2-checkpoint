
import openpyxl
import re

def extract_time(line):
    match = re.search(r'time=([\d.]+)', line)
    return float(match.group(1)) if match else None

# 创建一个新的 Excel 工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# 在 Excel 中写入标题行
sheet['A1'] = 'epoch'
sheet['B1'] = 'time'

# 逐行读取并处理文本文件
row_num = 2
count=1
with open('result_re_all2.txt', 'r') as file:
    for line in file:
        time_data = extract_time(line)
        if time_data is not None:
            # 在 Excel 中写入数据
            sheet.cell(row=row_num, column=1, value=20 * count)
            sheet.cell(row=row_num, column=2, value=time_data)
            count += 1
            row_num += 1

# 保存 Excel 文件
workbook.save('output.xlsx')
